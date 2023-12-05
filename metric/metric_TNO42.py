import numpy as np
from PIL import Image
from Metric import *
from natsort import natsorted
from tqdm import tqdm
import os
import statistics
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

def write_excel(excel_name='metric.xlsx', worksheet_name='VIF', column_index=0, data=None):
    try:
        workbook = load_workbook(excel_name)
    except FileNotFoundError:
    # 文件不存在，创建新的 Workbook
        workbook = Workbook()

    # 获取或创建一个工作表
    if worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
    else:
        worksheet = workbook.create_sheet(title=worksheet_name)

    # 在指定列中插入数据
    column = get_column_letter(column_index + 1)
    for i, value in enumerate(data):
        cell = worksheet[column + str(i+1)]
        cell.value = value

    # 保存文件
    workbook.save(excel_name)

def evaluation_one(ir_name, vi_name, f_name):
    f_img = Image.open(f_name).convert('L')
    ir_img = Image.open(ir_name).convert('L')
    vi_img = Image.open(vi_name).convert('L')

    f_img_int = np.array(f_img).astype(np.int32)
    f_img_double = np.array(f_img).astype(np.float32)

    ir_img_int = np.array(ir_img).astype(np.int32)
    ir_img_double = np.array(ir_img).astype(np.float32)

    vi_img_int = np.array(vi_img).astype(np.int32)
    vi_img_double = np.array(vi_img).astype(np.float32)

    EN = EN_function(f_img_int)
    MI = MI_function(ir_img_int, vi_img_int, f_img_int, gray_level=256)

    SF = SF_function(f_img_double)
    SD = SD_function(f_img_double)
    AG = AG_function(f_img_double)
    PSNR = PSNR_function(ir_img_double, vi_img_double, f_img_double)
    MSE = MSE_function(ir_img_double, vi_img_double, f_img_double)
    VIF = VIF_function(ir_img_double, vi_img_double, f_img_double)
    CC = CC_function(ir_img_double, vi_img_double, f_img_double)
    SCD = SCD_function(ir_img_double, vi_img_double, f_img_double)
    Qabf = Qabf_function(ir_img_double, vi_img_double, f_img_double)
    Nabf = Nabf_function(ir_img_double, vi_img_double, f_img_double)
    SSIM = SSIM_function(ir_img_double, vi_img_double, f_img_double)
    MS_SSIM = MS_SSIM_function(ir_img_double, vi_img_double, f_img_double)
    return EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM

if __name__ == '__main__':
    with_mean = True
    # dataroot = 'E:/task/fusion\RFN-NEST\images\TNO42'#测试集地址
    dataroot = r'E:\dataset\fusion\rgb\TNO42/'
    # results_root = r'E:\task\fusion\PIAFusion\result_image\TNO42'
    results_root = r'E:\task\fusion\PIAFusion\result_image\enw_mdw'
    #融合地址
    dataset = 'TNO42'
    ir_dir = os.path.join(dataroot, 'Inf')

    vi_dir = os.path.join(dataroot, 'Vis')
    # f_dir = os.path.join(results_root, dataset)
    f_dir = results_root
    # save_dir = r'E:\task\fusion\PIAFusion\result_image\TNO42'
    save_dir = r'E:\task\fusion\PIAFusion\result_image\enw_mdw'
    os.makedirs(save_dir, exist_ok=True)

    metric_save_name = os.path.join(save_dir, 'metric_{}.xlsx'.format(dataset))
    filelist = natsorted(os.listdir(ir_dir))

    Method_list = []
    for root, dirs, files in os.walk(save_dir):
        for dir_name in dirs:
            Method_list.append(dir_name)

    # Method_list = ['RFN-Nest_kd1', 'RFN-Nest_kd2', 'RFN-Nest_skfusion_5.0_0.5',
    #                'RFN-Nest_skfcafusion_6.0_3.0',
    #                'RFN-Nest_skfusion_6.0_3.0', 'RFN-Nest_skfusion_6.0_3.0_balance_loss']
    metric_list = ['','EN', 'MI', 'SF', 'AG', 'SD', 'CC',
                   'SCD', 'VIF', 'MSE', 'PSNR', 'Qabf',
                   'Nabf', 'SSIM', 'MS_SSIM']

    for i, Method in enumerate(Method_list):
        EN_list = []
        MI_list = []
        SF_list = []
        AG_list = []
        SD_list = []
        CC_list = []
        SCD_list = []
        VIF_list = []
        MSE_list = []
        PSNR_list = []
        Qabf_list = []
        Nabf_list = []
        SSIM_list = []
        MS_SSIM_list = []
        filename_list = ['']
        sub_f_dir = os.path.join(f_dir, Method)
        eval_bar = tqdm(filelist)
        for _, item in enumerate(eval_bar):
            ir_name = os.path.join(ir_dir, item)
            vi_name = os.path.join(vi_dir, item)
            f_name = os.path.join(sub_f_dir, item)
            EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM = evaluation_one(ir_name, vi_name, f_name)
            EN_list.append(EN)
            MI_list.append(MI)
            SF_list.append(SF)
            AG_list.append(AG)
            SD_list.append(SD)
            CC_list.append(CC)
            SCD_list.append(SCD)
            VIF_list.append(VIF)
            MSE_list.append(MSE)
            PSNR_list.append(PSNR)
            Qabf_list.append(Qabf)
            Nabf_list.append(Nabf)
            SSIM_list.append(SSIM)
            MS_SSIM_list.append(MS_SSIM)
            filename_list.append(item)
            eval_bar.set_description("{} | {}".format(Method, item))
        if with_mean:
        # 添加均值
            EN_list.append(np.mean(EN_list))
            MI_list.append(np.mean(MI_list))
            SF_list.append(np.mean(SF_list))
            AG_list.append(np.mean(AG_list))
            SD_list.append(np.mean(SD_list))
            CC_list.append(np.mean(CC_list))
            SCD_list.append(np.mean(SCD_list))
            VIF_list.append(np.mean(VIF_list))
            MSE_list.append(np.mean(MSE_list))
            PSNR_list.append(np.mean(PSNR_list))
            Qabf_list.append(np.mean(Qabf_list))
            Nabf_list.append(np.mean(Nabf_list))
            SSIM_list.append(np.mean(SSIM_list))
            MS_SSIM_list.append(np.mean(MS_SSIM_list))
            filename_list.append('mean')

            ## 添加标准差
            EN_list.append(np.std(EN_list))
            MI_list.append(np.std(MI_list))
            SF_list.append(np.std(SF_list))
            AG_list.append(np.std(AG_list))
            SD_list.append(np.std(SD_list))
            CC_list.append(np.std(CC_list[:-1]))
            SCD_list.append(np.std(SCD_list))
            VIF_list.append(np.std(VIF_list))
            MSE_list.append(np.std(MSE_list))
            PSNR_list.append(np.std(PSNR_list))
            Qabf_list.append(np.std(Qabf_list))
            Nabf_list.append(np.std(Nabf_list))
            SSIM_list.append(np.std(SSIM_list))
            MS_SSIM_list.append(np.std(MS_SSIM_list))
            filename_list.append('std')

        ## 保留三位小数
        EN_list = [round(x, 3) for x in EN_list]
        MI_list = [round(x, 3) for x in MI_list]
        SF_list = [round(x, 3) for x in SF_list]
        AG_list = [round(x, 3) for x in AG_list]
        SD_list = [round(x, 3) for x in SD_list]
        CC_list = [round(x, 3) for x in CC_list]
        SCD_list = [round(x, 3) for x in SCD_list]
        VIF_list = [round(x, 3) for x in VIF_list]
        MSE_list = [round(x, 3) for x in MSE_list]
        PSNR_list = [round(x, 3) for x in PSNR_list]
        Qabf_list = [round(x, 3) for x in Qabf_list]
        Nabf_list = [round(x, 3) for x in Nabf_list]
        SSIM_list = [round(x, 3) for x in SSIM_list]
        MS_SSIM_list = [round(x, 3) for x in MS_SSIM_list]

        EN_list.insert(0, '{}'.format(Method))
        MI_list.insert(0, '{}'.format(Method))
        SF_list.insert(0, '{}'.format(Method))
        AG_list.insert(0, '{}'.format(Method))
        SD_list.insert(0, '{}'.format(Method))
        CC_list.insert(0, '{}'.format(Method))
        SCD_list.insert(0, '{}'.format(Method))
        VIF_list.insert(0, '{}'.format(Method))
        MSE_list.insert(0, '{}'.format(Method))
        PSNR_list.insert(0, '{}'.format(Method))
        Qabf_list.insert(0, '{}'.format(Method))
        Nabf_list.insert(0, '{}'.format(Method))
        SSIM_list.insert(0, '{}'.format(Method))
        MS_SSIM_list.insert(0, '{}'.format(Method))

        mean_list = [EN_list[0],  EN_list[-2], MI_list[-2], SF_list[-2],
                     AG_list[-2], SD_list[-2], CC_list[-2],
                     SCD_list[-2], VIF_list[-2], MSE_list[-2],
                     PSNR_list[-2], Qabf_list[-2], Nabf_list[-2],
                     SSIM_list[-2], MS_SSIM_list[-2]]
        MI_list_mean = [MI_list[0], MI_list[-2]]
        SF_list_mean = [SF_list[0], SF_list[-2]]
        AG_list_mean = [AG_list[0], AG_list[-2]]
        SD_list_mean = [SD_list[0], SD_list[-2]]
        CC_list_mean = [CC_list[0], CC_list[-2]]
        SCD_list_mean = [SCD_list[0], SCD_list[-2]]
        VIF_list_mean = [VIF_list[0], VIF_list[-2]]
        MSE_list_mean = [MSE_list[0], MSE_list[-2]]
        PSNR_list_mean = [PSNR_list[0], PSNR_list[-2]]
        Qabf_list_mean = [Qabf_list[0], Qabf_list[-2]]
        Nabf_list_mean = [Nabf_list[0], Nabf_list[-2]]
        SSIM_list_mean = [SSIM_list[0], SSIM_list[-2]]
        MS_SSIM_list_mean = [MS_SSIM_list[0], MS_SSIM_list[-2]]

        if i == 0:
            write_excel(metric_save_name, "total_metric", 0, metric_list)
            write_excel(metric_save_name, 'EN', 0, filename_list)
            write_excel(metric_save_name, "MI", 0, filename_list)
            write_excel(metric_save_name, "SF", 0, filename_list)
            write_excel(metric_save_name, "AG", 0, filename_list)
            write_excel(metric_save_name, "SD", 0, filename_list)
            write_excel(metric_save_name, "CC", 0, filename_list)
            write_excel(metric_save_name, "SCD", 0, filename_list)
            write_excel(metric_save_name, "VIF", 0, filename_list)
            write_excel(metric_save_name, "MSE", 0, filename_list)
            write_excel(metric_save_name, "PSNR", 0, filename_list)
            write_excel(metric_save_name, "Qabf", 0, filename_list)
            write_excel(metric_save_name, "Nabf", 0, filename_list)
            write_excel(metric_save_name, "SSIM", 0, filename_list)
            write_excel(metric_save_name, "MS_SSIM", 0, filename_list)

        write_excel(metric_save_name, 'total_metric', i + 1, mean_list)
        # write_excel(metric_save_name, 'total_metric', i + 1, MI_list_mean)
        # write_excel(metric_save_name, 'total_metric', i + 3, SF_list_mean)
        # write_excel(metric_save_name, 'total_metric', i + 4, AG_list_mean)
        # write_excel(metric_save_name, 'total_metric', i + 5, SD_list_mean)
        # write_excel(metric_save_name, 'total_metric', i + 6, CC_list_mean)
        # write_excel(metric_save_name, 'total_metric', i + 7, SCD_list_mean)
        # write_excel(metric_save_name, 'total_metric', i + 8, VIF_list_mean)
        # write_excel(metric_save_name, 'total_metric', i + 9, MSE_list_mean)
        # write_excel(metric_save_name, 'total_metric', i + 10, PSNR_list_mean)
        # write_excel(metric_save_name, 'total_metric', i + 11, Qabf_list_mean)
        # write_excel(metric_save_name, 'total_metric', i + 12, Nabf_list_mean)
        # write_excel(metric_save_name, 'total_metric', i + 13, SSIM_list_mean)
        # write_excel(metric_save_name, 'total_metric', i + 14, MS_SSIM_list_mean)

        write_excel(metric_save_name, 'EN', i + 1, EN_list)
        write_excel(metric_save_name, 'MI', i + 1, MI_list)
        write_excel(metric_save_name, 'SF', i + 1, SF_list)
        write_excel(metric_save_name, 'AG', i + 1, AG_list)
        write_excel(metric_save_name, 'SD', i + 1, SD_list)
        write_excel(metric_save_name, 'CC', i + 1, CC_list)
        write_excel(metric_save_name, 'SCD', i + 1, SCD_list)
        write_excel(metric_save_name, 'VIF', i + 1, VIF_list)
        write_excel(metric_save_name, 'MSE', i + 1, MSE_list)
        write_excel(metric_save_name, 'PSNR', i + 1, PSNR_list)
        write_excel(metric_save_name, 'Qabf', i + 1, Qabf_list)
        write_excel(metric_save_name, 'Nabf', i + 1, Nabf_list)
        write_excel(metric_save_name, 'SSIM', i + 1, SSIM_list)
        write_excel(metric_save_name, 'MS_SSIM', i + 1, MS_SSIM_list)